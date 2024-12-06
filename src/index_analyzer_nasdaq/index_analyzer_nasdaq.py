import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

# Step 1: Define the range of years
years = list(range(1990, 2025))

# Fetch NASDAQ-100 constituents from Wikipedia
nasdaq_100_url = 'https://en.wikipedia.org/wiki/NASDAQ-100'
tables = pd.read_html(nasdaq_100_url)

# Initialize variables
nasdaq_100_df = None
ticker_column = None
name_column = None

# Find the correct table and columns
for idx, table in enumerate(tables):
    # Possible column names for tickers and company names
    possible_ticker_columns = ['Ticker', 'Ticker symbol', 'Symbol']
    possible_name_columns = ['Company', 'Security', 'Name']
    columns = table.columns.tolist()
    if any(col in possible_ticker_columns for col in columns) and any(col in possible_name_columns for col in columns):
        nasdaq_100_df = table
        for col in columns:
            if col in possible_ticker_columns:
                ticker_column = col
            if col in possible_name_columns:
                name_column = col
        break

if nasdaq_100_df is None or ticker_column is None or name_column is None:
    raise ValueError("Could not find NASDAQ-100 table or required columns on Wikipedia page.")

# Extract the list of tickers and company names
tickers = nasdaq_100_df[ticker_column].tolist()
company_names = nasdaq_100_df[name_column].tolist()

# Remove any non-string tickers and company names
tickers = [str(ticker) for ticker in tickers if isinstance(ticker, str)]
company_names = [str(name) for name in company_names if isinstance(name, str)]

# Remove any periods from tickers (e.g., BRK.B -> BRK-B)
tickers = [ticker.replace('.', '-') for ticker in tickers]

# Create a dictionary mapping tickers to company names
ticker_to_name = dict(zip(tickers, company_names))

# Create dictionaries to store stock data and market cap for each company
data = {}
market_caps = {}  # New dictionary to store market caps

# Step 2: Loop through each ticker and gather annual data and market cap
for ticker in tickers:
    print(f"Processing {ticker}...")
    stock = yf.Ticker(ticker)
    # Get historical data
    hist = stock.history(period="max")
    if hist.empty:
        print(f"No historical data for {ticker}. Skipping.")
        continue  # Skip if no historical data is available
    hist['Year'] = hist.index.year
    # Calculate year-over-year growth percentage
    yearly_close = hist['Close'].resample('Y').last()
    yearly_data = yearly_close.pct_change() * 100
    yearly_data.index = yearly_data.index.year  # Use year as index
    yearly_data = yearly_data.round(1)  # Round to one decimal place
    data[ticker] = yearly_data.reindex(years)  # Reindex to match years range

    # Fetch market cap using stock.fast_info
    try:
        market_cap = stock.fast_info['market_cap']
    except Exception as e:
        market_cap = None
    market_caps[ticker] = market_cap

# Step 3: Create DataFrame from dictionary
df = pd.DataFrame(data).transpose()
df.index.name = 'Ticker'

# Add Market Cap column (numeric values)
df['Market Cap'] = df.index.map(market_caps)

# **Create a temporary numeric column for sorting**
df['Market Cap Numeric'] = df['Market Cap']

# **Sort the DataFrame by Market Cap Numeric in descending order**
df.sort_values(by='Market Cap Numeric', ascending=False, inplace=True)

# **Format the Market Cap values for readability**
def format_market_cap(market_cap):
    if pd.isna(market_cap):
        return ''
    elif market_cap >= 1e12:
        return f'{market_cap / 1e12:.1f}T'  # Trillions
    elif market_cap >= 1e9:
        return f'{market_cap / 1e9:.1f}B'   # Billions
    elif market_cap >= 1e6:
        return f'{market_cap / 1e6:.1f}M'   # Millions
    else:
        return str(market_cap)

df['Market Cap'] = df['Market Cap'].apply(format_market_cap)

# **Drop the temporary numeric column**
df.drop(columns=['Market Cap Numeric'], inplace=True)

# Round the DataFrame values to one decimal place (in case any values were missed)
df = df.round(1)

# Reset the index to turn the ticker symbol into a column
df.reset_index(inplace=True)

# Add the company name column
df['Company Name'] = df['Ticker'].map(ticker_to_name)

# Reorder columns to have Ticker, Company Name, Market Cap, then the years
cols = ['Ticker', 'Company Name', 'Market Cap'] + [col for col in df.columns if col not in ['Ticker', 'Company Name', 'Market Cap']]
df = df[cols]

# Step 4: Export DataFrame to Excel with conditional formatting
file_name = "NASDAQ_100_Annual_Growth.xlsx"

with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Growth", index=False)
    workbook = writer.book
    sheet = writer.sheets["Growth"]
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define alignment
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Define fills for green shades
    green_fills = [
        PatternFill(start_color="e2f0d9", end_color="e2f0d9", fill_type="solid"),  # Very light green
        PatternFill(start_color="c6e0b4", end_color="c6e0b4", fill_type="solid"),  # Light green
        PatternFill(start_color="a9d08e", end_color="a9d08e", fill_type="solid"),  # Medium green
        PatternFill(start_color="70ad47", end_color="70ad47", fill_type="solid"),  # Dark green
        PatternFill(start_color="548235", end_color="548235", fill_type="solid")   # Darkest green
    ]
    
    # Define fills for red shades
    red_fills = [
        PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid"),  # Very light red
        PatternFill(start_color="ea9999", end_color="ea9999", fill_type="solid"),  # Light red
        PatternFill(start_color="e06666", end_color="e06666", fill_type="solid"),  # Medium red
        PatternFill(start_color="cc0000", end_color="cc0000", fill_type="solid"),  # Dark red
        PatternFill(start_color="990000", end_color="990000", fill_type="solid")   # Darkest red
    ]
    
    # Apply borders and alignment to all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    # Identify the data columns (starting from the fourth column, since first three are Ticker, Company Name, and Market Cap)
    data_col_start = 4  # Updated to 4 because of the new 'Market Cap' column
    data_col_end = sheet.max_column

    # Iterate over data cells to apply fills
    for row in range(2, sheet.max_row + 1):  # Starting from row 2 to skip header
        for col in range(data_col_start, data_col_end + 1):
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            if pd.notna(value):
                try:
                    numeric_value = float(value)
                    cell.number_format = '0.0'  # Set number format to one decimal place
                    
                    # Determine the fill color based on value
                    if numeric_value > 0:
                        if 0 < numeric_value <= 5:
                            fill = green_fills[0]  # Very light green
                        elif 5 < numeric_value <= 10:
                            fill = green_fills[1]  # Light green
                        elif 10 < numeric_value <= 20:
                            fill = green_fills[2]  # Medium green
                        elif 20 < numeric_value <= 50:
                            fill = green_fills[3]  # Dark green
                        elif numeric_value > 50:
                            fill = green_fills[4]  # Darkest green
                    elif numeric_value < 0:
                        if -5 <= numeric_value < 0:
                            fill = red_fills[0]  # Very light red
                        elif -10 <= numeric_value < -5:
                            fill = red_fills[1]  # Light red
                        elif -20 <= numeric_value < -10:
                            fill = red_fills[2]  # Medium red
                        elif -50 <= numeric_value < -20:
                            fill = red_fills[3]  # Dark red
                        elif numeric_value < -50:
                            fill = red_fills[4]  # Darkest red
                    else:
                        fill = None  # Zero change
                    if fill:
                        cell.fill = fill
                except ValueError:
                    pass  # Skip if the value cannot be converted to float

print(f"Data has been saved to {file_name}")
