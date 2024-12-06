import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Step 1: Define the range of years
years = list(range(1990, 2025))

# Hardcoded list of ticker symbols
tickers = [
    'IONQ', 'MSTR', 'QBTS', 'RGTI', 'COHR', 'RDNT', '2357.TW', 'MRVL', 'ACN', 'IBM',
    'TXN', 'HPE', 'NTTYY', '6503.T', 'KPN', 'GOOGL', 'ORAN', 'ABB', 'CDNS', '2454.TW',
    'MSFT', 'SNPS', 'NVEC', 'ADI', 'EADSY', 'ON', 'BIDU', 'AMD', 'TDC', '3661.TW',
    'CRUS', 'MKSI', 'LSCC', 'SYNA', 'KLAC', 'NXPI', 'INTC', 'REY.MI', 'NOC', 'WIT',
    'LMT', 'BABA', 'BAH', 'TSEM', 'NOK', 'NVDA', 'HTHIY', '9613.T', 'RTX', 'FJTSY',
    'HON', 'TSM', 'JNPR', '6701.T'
]

# Remove any periods from tickers (e.g., BRK.B -> BRK-B) for yfinance compatibility
tickers = [ticker.replace('.', '-') for ticker in tickers]

# Create dictionaries to store stock data and company names
data = {}
ticker_to_name = {}

# Step 2: Loop through each ticker, retrieve company name, and gather annual data
valid_tickers = []  # List to keep track of tickers with valid data
for ticker in tickers:
    print(f"Processing {ticker}...")
    stock = yf.Ticker(ticker)
    # Retrieve company name
    info = stock.info
    company_name = info.get('shortName') or info.get('longName') or 'N/A'
    ticker_to_name[ticker] = company_name
    # Get historical data
    hist = stock.history(period="max")
    if hist.empty:
        print(f"No historical data for {ticker}. Skipping.")
        continue  # Skip if no historical data is available
    hist['Year'] = hist.index.year
    # Calculate year-over-year growth percentage
    yearly_close = hist['Close'].resample('Y').last()
    yearly_data = yearly_close.pct_change() * 100
    yearly_data.index = yearly_data.index.year  # Use year as index
    yearly_data = yearly_data.round(1)  # Round to one decimal place
    data[ticker] = yearly_data.reindex(years)  # Reindex to match years range
    valid_tickers.append(ticker)  # Add to valid tickers list

# Step 3: Create DataFrame from dictionary
df = pd.DataFrame(data).transpose()
df.index.name = 'Ticker'

# Round the DataFrame values to one decimal place (in case any values were missed)
df = df.round(1)

# Reset the index to turn the ticker symbol into a column
df.reset_index(inplace=True)

# Add the company name column
df['Company Name'] = df['Ticker'].map(ticker_to_name)

# Reorder columns to have Ticker, Company Name, then the years
cols = ['Ticker', 'Company Name'] + [col for col in df.columns if col not in ['Ticker', 'Company Name']]
df = df[cols]

# Calculate the average annual growth rate, treating the companies as an equally weighted ETF
# Exclude companies without data in a given year
avg_growth = df.set_index(['Ticker', 'Company Name']).mean(axis=0, skipna=True).round(1)
avg_growth = avg_growth.to_frame().T
avg_growth.insert(0, 'Company Name', 'Weighted Mean (ETF)')
avg_growth.insert(0, 'Ticker', '')

# Append a blank row and then the average growth row to the DataFrame
blank_row = pd.DataFrame([['', ''] + ['' for _ in range(len(years))]], columns=df.columns)
df = pd.concat([df, blank_row, avg_growth], ignore_index=True)

# Step 4: Export DataFrame to Excel with conditional formatting
file_name = "Stock_Annual_Growth.xlsx"
with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Growth", index=False)
    workbook = writer.book
    sheet = writer.sheets["Growth"]
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define alignment
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Define fills for green shades
    green_fills = [
        PatternFill(start_color="e2f0d9", end_color="e2f0d9", fill_type="solid"),  # Very light green
        PatternFill(start_color="c6e0b4", end_color="c6e0b4", fill_type="solid"),  # Light green
        PatternFill(start_color="a9d08e", end_color="a9d08e", fill_type="solid"),  # Medium green
        PatternFill(start_color="70ad47", end_color="70ad47", fill_type="solid"),  # Dark green
        PatternFill(start_color="548235", end_color="548235", fill_type="solid")   # Darkest green
    ]
    
    # Define fills for red shades
    red_fills = [
        PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid"),  # Very light red
        PatternFill(start_color="ea9999", end_color="ea9999", fill_type="solid"),  # Light red
        PatternFill(start_color="e06666", end_color="e06666", fill_type="solid"),  # Medium red
        PatternFill(start_color="cc0000", end_color="cc0000", fill_type="solid"),  # Dark red
        PatternFill(start_color="990000", end_color="990000", fill_type="solid")   # Darkest red
    ]
    
    # Apply borders and alignment to all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    # Set the width of the 'Company Name' column to be wider
    company_name_col = 2  # Assuming 'Company Name' is the second column (B)
    company_name_col_letter = get_column_letter(company_name_col)
    sheet.column_dimensions[company_name_col_letter].width = 30  # Adjust width as needed

    # Identify the data columns (starting from the third column)
    data_col_start = 3  # Data starts from the third column
    data_col_end = sheet.max_column

    # Iterate over data cells to apply fills
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Starting from row 2 to skip header
        # Skip the blank row before the average growth row
        if all(cell.value == '' or cell.value is None for cell in row):
            continue
        for col in range(data_col_start, data_col_end + 1):
            cell = sheet.cell(row=row_idx, column=col)
            value = cell.value
            if pd.notna(value) and isinstance(value, (int, float)):
                numeric_value = float(value)
                cell.number_format = '0.0'  # Set number format to one decimal place
                
                # Determine the fill color based on value
                if numeric_value > 0:
                    if 0 < numeric_value <= 5:
                        fill = green_fills[0]  # Very light green
                    elif 5 < numeric_value <= 10:
                        fill = green_fills[1]  # Light green
                    elif 10 < numeric_value <= 20:
                        fill = green_fills[2]  # Medium green
                    elif 20 < numeric_value <= 50:
                        fill = green_fills[3]  # Dark green
                    elif numeric_value > 50:
                        fill = green_fills[4]  # Darkest green
                elif numeric_value < 0:
                    if -5 <= numeric_value < 0:
                        fill = red_fills[0]  # Very light red
                    elif -10 <= numeric_value < -5:
                        fill = red_fills[1]  # Light red
                    elif -20 <= numeric_value < -10:
                        fill = red_fills[2]  # Medium red
                    elif -50 <= numeric_value < -20:
                        fill = red_fills[3]  # Dark red
                    elif numeric_value < -50:
                        fill = red_fills[4]  # Darkest red
                else:
                    fill = None  # Zero change
                if fill:
                    cell.fill = fill
            else:
                cell.number_format = '0.0'  # Ensure number format is consistent

print(f"Data has been saved to {file_name}")
